import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import Font, PatternFill
import re

st.title("Diamond Tool")

# ================= SIZE GROUP FUNCTION =================

def get_size_grp(cts):
    if pd.isna(cts):
        return ""

    cts = float(cts)

    if 0.30 <= cts <= 0.49:
        return "0.30 - 0.49"
    elif 0.50 <= cts <= 0.59:
        return "0.50 - 0.59"
    elif 0.60 <= cts <= 0.69:
        return "0.60 - 0.69"
    elif 0.70 <= cts <= 0.79:
        return "0.70 - 0.79"
    elif 0.80 <= cts <= 0.89:
        return "0.80 - 0.89"
    elif 0.90 <= cts <= 0.99:
        return "0.90 - 0.99"
    elif 1.00 <= cts <= 1.05:
        return "1.00 - 1.05"
    elif 1.06 <= cts <= 1.10:
        return "1.06 - 1.10"
    elif 1.11 <= cts <= 1.19:
        return "1.11 - 1.19"
    elif 1.20 <= cts <= 1.29:
        return "1.20 - 1.29"
    elif 1.30 <= cts <= 1.39:
        return "1.30 - 1.39"
    elif 1.40 <= cts <= 1.49:
        return "1.40 - 1.49"
    elif 1.50 <= cts <= 1.55:
        return "1.50 - 1.55"
    elif 1.56 <= cts <= 1.59:
        return "1.56 - 1.59"
    elif 1.60 <= cts <= 1.69:
        return "1.60 - 1.69"
    elif 1.70 <= cts <= 1.79:
        return "1.70 - 1.79"
    elif 1.80 <= cts <= 1.89:
        return "1.80 - 1.89"
    elif 1.90 <= cts <= 1.99:
        return "1.90 - 1.99"
    elif 2.00 <= cts <= 2.05:
        return "2.00 - 2.05"
    elif 2.06 <= cts <= 2.10:
        return "2.06 - 2.10"
    elif 2.11 <= cts <= 2.19:
        return "2.11 - 2.19"
    elif 2.20 <= cts <= 2.29:
        return "2.20 - 2.29"
    elif 2.30 <= cts <= 2.39:
        return "2.30 - 2.39"
    elif 2.40 <= cts <= 2.49:
        return "2.40 - 2.49"
    elif 2.50 <= cts <= 2.55:
        return "2.50 - 2.55"
    elif 2.56 <= cts <= 2.59:
        return "2.56 - 2.59"
    elif 2.60 <= cts <= 2.69:
        return "2.60 - 2.69"
    elif 2.70 <= cts <= 2.79:
        return "2.70 - 2.79"
    elif 2.80 <= cts <= 2.89:
        return "2.80 - 2.89"
    elif 2.90 <= cts <= 2.99:
        return "2.90 - 2.99"
    elif 3.00 <= cts <= 3.05:
        return "3.00 - 3.05"
    elif 3.06 <= cts <= 3.10:
        return "3.06 - 3.10"
    elif 3.11 <= cts <= 3.19:
        return "3.11 - 3.19"
    elif 3.20 <= cts <= 3.29:
        return "3.20 - 3.29"
    elif 3.30 <= cts <= 3.39:
        return "3.30 - 3.39"
    elif 3.40 <= cts <= 3.49:
        return "3.40 - 3.49"
    elif 3.50 <= cts <= 3.55:
        return "3.50 - 3.55"
    elif 3.56 <= cts <= 3.59:
        return "3.56 - 3.59"
    elif 3.60 <= cts <= 3.69:
        return "3.60 - 3.69"
    elif 3.70 <= cts <= 3.79:
        return "3.70 - 3.79"
    elif 3.80 <= cts <= 3.89:
        return "3.80 - 3.89"
    elif 3.90 <= cts <= 3.99:
        return "3.90 - 3.99"
    elif 4.00 <= cts <= 4.10:
        return "4.00 - 4.10"
    elif 4.11 <= cts <= 4.49:
        return "4.11 - 4.49"
    elif 4.50 <= cts <= 4.59:
        return "4.50 - 4.59"
    elif 4.60 <= cts <= 4.99:
        return "4.60 - 4.99"
    elif 5.00 <= cts <= 5.10:
        return "5.00 - 5.10"
    elif 5.11 <= cts <= 5.49:
        return "5.11 - 5.49"
    elif 5.50 <= cts <= 5.59:
        return "5.50 - 5.59"
    elif 5.60 <= cts <= 5.99:
        return "5.60 - 5.99"
    elif 6.00 <= cts <= 6.10:
        return "6.00 - 6.10"
    elif 6.11 <= cts <= 6.49:
        return "6.11 - 6.49"
    elif 6.50 <= cts <= 6.59:
        return "6.50 - 6.59"
    elif 6.60 <= cts <= 6.99:
        return "6.60 - 6.99"
    elif 7.00 <= cts <= 7.99:
        return "7.00 - 7.99"
    elif 8.00 <= cts <= 8.99:
        return "8.00 - 8.99"
    else:
        return ""

# ================= MATCHING PAIR FUNCTIONS =================

pair_fill = PatternFill(
    fill_type="solid",
    start_color="7030A0",   # Dark Purple
    end_color="7030A0"
)

def get_pair_base(lot):

    lot = str(lot).strip().upper()
    if re.match(r"^.+[A-Z]$", lot):
        return lot[:-1]

    return None




# ================= FILE UPLOAD =================

cost_file = st.file_uploader("Upload Cost File", type=["xlsx"])
panding_file = st.file_uploader("Upload Pending File", type=["xlsx"])
lab_file = st.file_uploader("Upload Lab File", type=["xls", "xlsx"])

if cost_file and panding_file and lab_file:

    # READ FILES
    cost = pd.read_excel(cost_file)
    panding = pd.read_excel(panding_file)

    if lab_file.name.endswith(".xls"):
        lab = pd.read_excel(lab_file, header=2, engine="xlrd")
    else:
        lab = pd.read_excel(lab_file, header=2, engine="openpyxl")

    # CLEAN COLUMN NAMES
    cost.columns = cost.columns.str.strip()
    panding.columns = panding.columns.str.strip()
    lab.columns = lab.columns.str.strip()

    # COST FILE REQUIRED COLUMNS
    cost = cost[[
        "Lot #",
        "Shape",
        "Color",
        "Clarity",
        "Cts.",
        "GIA #",
        "Lab",
        "Quality",
        "Price / Cts",
        "Cost / Cts.",
        "Rapnet Note"
    ]]

    # LAB FILTER
    cost = cost[cost["Lab"].isin(["GIA", "IGI", "GCAL"])]

    # COLOR FILTER
    valid_colors = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    cost["Color"] = cost["Color"].astype(str).str.strip()
    cost = cost[cost["Color"].isin(valid_colors)]

    # REMOVE VP SERIES
    cost["Lot #"] = cost["Lot #"].astype(str).str.strip()
    cost = cost[
        ~cost["Lot #"].str.upper().str.startswith("VP")
    ]

    # QUALITY FIX
    cost["Quality"] = cost["Quality"].fillna("").astype(str).str.strip()
    cost["Rapnet Note"] = cost["Rapnet Note"].fillna("").astype(str).str.upper()

    cost["Quality"] = cost["Quality"].replace(
        ["Blank", "blank", "BLANK", "nan", "NaN"],
        ""
    )

    cost.loc[
        (cost["Quality"] == "") &
        (cost["Rapnet Note"].str.contains("CVD", na=False)),
        "Quality"
    ] = "CVD"

    cost.loc[
        (cost["Quality"] == "") &
        (cost["Rapnet Note"].str.contains("HPHT", na=False)),
        "Quality"
    ] = "HPHT"

    # PENDING FILE FIX
    panding["Customer"] = (
        panding["Customer"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    panding["Status"] = (
        panding["Status"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    panding.loc[
        (
            (panding["Customer"] == "GOODS IN TRANSIT FROM OVERSEAS") |
            (panding["Customer"] == "GOODS IN OFFICE - PARCEL PAPERS BEING MADE")
        ) &
        (panding["Status"].str.upper() == "ONMEMO"),
        "Status"
    ] = "Inhand"

    # ================= FIX LOT MATCH =================
    cost["Lot #"] = (
        cost["Lot #"]
        .astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
        .str.upper()
    )

    panding["Lot #"] = (
        panding["Lot #"]
        .astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
        .str.upper()
    )

    # ================= MERGE STATUS =================
    panding = panding[["Lot #", "Status"]]
    cost = cost.merge(panding, on="Lot #", how="left")

    # LAB FILE CLEAN
    stock_col = [c for c in lab.columns if "stock" in c.lower()][0]
    days_col = [c for c in lab.columns if "old" in c.lower()][0]

    lab = lab[[stock_col, days_col]]

    lab = lab.rename(columns={
        stock_col: "Lot #",
        days_col: "No of Days"
    })

    # MERGE LAB
    cost = cost.merge(lab, on="Lot #", how="left")

    # NO OF DAYS FIX
    cost["No of Days"] = pd.to_numeric(cost["No of Days"], errors="coerce")

    cost.loc[
        (
            cost["Lot #"].str.upper().str.startswith(("DM", "DC"))
        ) &
        (
            cost["No of Days"] == 0
        ),
        "No of Days"
    ] = np.nan

    # SIZE GROUP
    cost["Cts."] = pd.to_numeric(cost["Cts."], errors="coerce")
    cost["Size Grp"] = cost["Cts."].apply(get_size_grp)

    # EXTRA HEADER COLUMNS ONLY
    cost["UPDATED PRICE"] = ""
    cost["DIFFERENCE"] = ""
    cost["Cost Amt"] = ""
    cost["Sale Amt"] = ""
    cost["Differance"] = ""

    # FINAL FORMAT
    cost = cost[[
        "Lot #",
        "Status",
        "Shape",
        "Color",
        "Clarity",
        "Cts.",
        "Size Grp",
        "No of Days",
        "Price / Cts",
        "Cost / Cts.",
        "GIA #",
        "Lab",
        "Quality",
        "UPDATED PRICE",
        "DIFFERENCE",
        "Cost Amt",
        "Sale Amt",
        "Differance"
    ]]
    # ================= FIND MATCHING PAIRS =================
    pair_groups = {}
    for lot in cost["Lot #"]:
         base = get_pair_base(lot)
         if base:
             pair_groups.setdefault(base, []).append(str(lot).upper())
    matched_pairs = {
        base
        for base, lots in pair_groups.items()
        if len(lots) >= 2
    }


    # OUTPUT
    st.success("Processing Completed Successfully ✅")

    total_diamond = len(cost)
    st.markdown(f"## Total Diamonds: {total_diamond}")
    st.markdown("---")

    st.dataframe(cost)

    # DOWNLOAD EXCEL
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        cost.to_excel(writer, index=False, sheet_name="Final Output")

        worksheet = writer.sheets["Final Output"]
        # ================= HIGHLIGHT MATCHING PAIRS =================
        for row in range(2, len(cost) + 2):
            lot = str(worksheet.cell(row=row, column=1).value).strip().upper()
            base = get_pair_base(lot)
            if base in matched_pairs:
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row, column=col).fill = pair_fill






        # Bold header
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # ================= ADD DIFFERENCE FORMULA =================
        diff_col_index = list(cost.columns).index("DIFFERENCE") + 1

        for row in range(2, len(cost) + 2):
            worksheet.cell(
                row=row,
                column=diff_col_index
            ).value = f"=-ROUND((J{row}-N{row})/J{row}%,2)"

    buffer.seek(0)

    st.download_button(
        label="Download Final Excel File",
        data=buffer,
        file_name="Final_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
